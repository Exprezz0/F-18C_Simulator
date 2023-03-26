

import os
import pandas as pd
from openpyxl import load_workbook
import sqlite3 

# Set the directory containing the PCB folders
script_path = os.path.dirname(os.path.abspath(__file__))
base_dir = os.path.dirname(script_path)
os.chdir(base_dir)
#pcb_dir = os.path.dirname()

# Load the Excel spreadsheet containing the PCB names and quantities
wb = load_workbook(filename='docs/OH_PCB Database.xlsx', read_only=True)
ws = wb.active
pcb_quantities = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    pcb_quantities[row[0]] = row[1]

# Loop through each PCB folder and consolidate the BOMs
master_bom = {}
for pcb_name, quantity in pcb_quantities.items():
    top_bom_file = os.path.join(base_dir, pcb_name, 'top_bom.csv')
    bottom_bom_file = os.path.join(base_dir, pcb_name, 'bottom_bom.csv')
    
    top_bom = pd.read_csv(top_bom_file)
    bottom_bom = pd.read_csv(bottom_bom_file)
    
    # Combine the top and bottom BOMs into a single list of components
    components = top_bom.append(bottom_bom, ignore_index=True)
    
    # Multiply the quantity of each component by the number of PCBs needed
    components['Quantity'] = components['Quantity'] * quantity
    
    # Add the multiplied components to the master BOM
    for i, row in components.iterrows():
        component = row['Component']
        quantity = row['Quantity']
        if component in master_bom:
            master_bom[component] += quantity
        else:
            master_bom[component] = quantity

# Write the master BOM to a CSV file
with open('master_bom.csv', 'w') as f:
    f.write('Component,Quantity\n')
    for component, quantity in master_bom.items():
        f.write(f'{component},{quantity}\n')
